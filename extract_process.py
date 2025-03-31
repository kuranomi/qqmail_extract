import pandas as pd
import numpy as np
import re
from gfile import GFile
import zipfile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import unicodedata
from send2trash import send2trash

words_pair = [
    [r'^(?=[\s\S]*jan)(?=[\s\S]*(?<!\w)box(?!\w))[\s\S]*$',"jan_box"],
    [r'^(?=[\s\S]*jan)(?=[\s\S]*(?<!\w)pcs|p(?!\w))[\s\S]*$', "jan_pcs"],
    [r'^(?=.*jan)(?!.*\b(?:box|pcs)\b).+$',"jan"],
    [r'ip',"ip"],
    [r'title', "ip"],
    [r'タイトル',"ip"],
    [r'品名',"name"],
    [r'受注単位',"standard"],
    [r'注文数',"standard"],
    [r'^(?=.*box)(?=.*入数).*$',"number_box"],
    [r'^(?=.*pcs)(?=.*入数).*$',"number_pcs"],
    [r'入数', "number"],
    [r'^(?=.*注締)(?=.*\日\b).+$',"cutoff_date"],
    [r'^(?=.*納品予定).+$',"release_date"],
    [r'^(?=.*発売日).+$',"release_date"],
    [r'税込', "price_with_tax"],
    [r'税抜',"price_without_tax"],
    [r'上代', "price"],
    [r'掛率',"discount"]
]


def is_gfile(l):
    lp = l.split(".")
    if "gigafile" in lp:
        return True
    return False

def image_extract(links, extract_dir):
    all_pic_path = []
    for link in links:
        if not is_gfile(link):
            continue 
        try:
            print("images downloading.....")
            filenames = GFile(link).download()
            if not filenames:
                print(f"invalid link: {link}")
                continue
            print(f"finish downloading {filenames}")
        except Exception:
            print(f"invalid link: {link}")
            continue
        try:
            for fn in filenames:
                with zipfile.ZipFile(fn, 'r') as zip_ref:
                    file_info_list = zip_ref.infolist()

                    print("pic extracting.....")
                    zip_ref.extractall(extract_dir)

                    pathes = [info.filename for info in file_info_list]
                    all_pic_path.extend(pathes)
                send2trash(fn)
        except zipfile.BadZipFile:
            print("Error: not valid zip file")
            return []
        except PermissionError:
            print("Error: Permission Denied")
            return []
    return all_pic_path


def find_keywords(key):
    if not type(key) is str:
        return None
    key = unicodedata.normalize('NFKC', key)
    key = key.lower()
    for k,v in words_pair:
        if re.search(k, key):
            return v
    return None


def find_info(row, info):
    key = None
    for col_idx, val in enumerate(row):
        if pd.isna(val):
            continue
        standard_kw = find_keywords(val)
        if standard_kw:
            key = standard_kw
        elif key:
            info[key] = val
            key = None
    return info


def check_table_type(df):
    need_content = {}
    for row_idx, row in df.iterrows():
        keywords = {}
        has_jan = False
        for col_idx, val in enumerate(row):
            if pd.isna(val):
                continue
            standard_kw = find_keywords(val)
            if standard_kw:
                if standard_kw.startswith("jan"):
                    has_jan = True
                keywords[standard_kw] = col_idx

        if has_jan:
            for jan in ["jan", "jan_box", "jan_pcs"]:
                if not jan in keywords.keys():
                    continue
                tmp_val = df.iloc[row_idx + 1, keywords.get(jan)]
                if str(tmp_val).isdigit() and len(str(tmp_val)) == 13:
                    return need_content, keywords, row_idx
            need_content = find_info(row, need_content)
            continue

        all_standard_column_name = list(keywords.keys()) + list(need_content.keys())
        if len(set(keywords.keys())) > 1 and len(set(all_standard_column_name)) >= 4 and \
                all(map(lambda s: s in all_standard_column_name, ["name", "jan"])):
            return need_content, keywords, row_idx

        need_content = find_info(row, need_content)
    return need_content, {}, 0


def get_data(xls, sheet_name):
    df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    df = df.dropna(axis=0, how="all")
    df.reset_index(inplace=True, drop=True)
    detail, columns, start_row = check_table_type(df)
    if columns:
        column_names = list(columns.keys())
        need_data = df.loc[start_row+1:, list(columns.values())].copy()
        need_data.columns = column_names
        if detail:
            for col, value in detail.items():
                if col in column_names:
                    # keep table value
                    continue
                need_data.loc[:, col] = value
    else:
        need_data = pd.DataFrame(detail)
    return clean_table(need_data)


def clean_table(df):
    columns = df.columns
    for cn in ["jan", "jan_box", "jan_pcs"]:
        if cn in columns:
            df[cn] = df[cn].apply(
                lambda x: str(int(x)) if str(x).isdigit() and len(str(x)) == 13 else np.nan
            )
    if not "jan" in columns and "jan_pcs" in columns:
        df.rename(columns={"jan_pcs": "jan"}, inplace=True)
    if not "number" in columns and "number_box" in columns:
        df.rename(columns={"number_box": "number"}, inplace=True)
    columns = df.columns
    if "discount" in columns:
        df["discount"] = pd.to_numeric(df["discount"], errors="coerce")
    try:
        if "cutoff_date" in columns:
            df["cutoff_date"] = pd.to_datetime(df["cutoff_date"])
            df["cutoff_date"] = df["cutoff_date"].dt.strftime("%Y-%m-%d")
        if "release_date" in columns:
            df["release_date"] = pd.to_datetime(df["release_date"])
            df["release_date"] = df["release_date"].dt.strftime("%Y-%m")
    except Exception:
        pass
    calculate = False
    if ("jan" in columns and "jan_box" in columns) or ("jan" in columns and "jan_pcs" in columns):
        calculate = True
    new_rows = []
    if calculate:
        for index, row in df.iterrows():
            new_rows.append(row)
            if "jan_box" in columns and pd.notna(row["jan_box"]):
                new_row = row.copy()
                new_row["jan"] = row["jan_box"]
                if "number" in columns and isinstance(row["number"], int):
                    for cn in columns:
                        if cn.startswith("price"):
                            new_p = row[cn] * row["number"]
                            new_row[cn] = new_p
                new_rows.append(new_row)
            if "jan_pcs" in columns and pd.notna(row["jan_pcs"]):
                new_row = row.copy()
                new_row["jan"] = row["jan_pcs"]
                if "number" in columns and isinstance(row["number"], int) and row["number"] > 0:
                    for cn in columns:
                        if cn.startswith("price"):
                            new_p = row[cn] / row["number"]
                            new_row[cn] = new_p
                            new_row["number"] = 1
                new_rows.append(new_row)

        new_df = pd.DataFrame(new_rows)
        if "jan_pcs" in columns:
            new_df = new_df.drop(columns=["jan_pcs"])
        if "jan_box" in columns:
            new_df = new_df.drop(columns=["jan_box"])
    else:
        new_df = df

    new_df = new_df.dropna(subset=["jan"])
    new_df = new_df.drop_duplicates(subset="jan", keep="first")
    fix_table = ["jan", "image", "name", "ip", "price_with_tax", "price_without_tax", "price", "discount", "number",
                 "standard", "cutoff_date", "release_date", "image_path"]
    for cn in fix_table:
        if cn not in columns:
            new_df[cn] = np.nan
    full_df = new_df[fix_table]
    full_df.reset_index(inplace=True, drop=True)
    return full_df

def add_image_path(image_path, return_df):
    jan = {}
    pattern = r'^.*?(\d{13}).*?\.jpg$'
    for path in image_path:
        is_img = re.match(pattern, path)
        if is_img:
            jn = is_img.group(1)
            jan[jn] = jan.get(jn, []) + [path]

    return_df["image_path"] = return_df["image_path"].astype(str)
    for row_idx, row in return_df.iterrows():
        jancode = str(row["jan"])
        if jancode and jan.get(jancode):
            image_link = jan[jancode][0]
            return_df.at[row_idx, "image_path"] = image_link
    return return_df

def inset_image_and_export(return_df, save_dir, image_path):
    df_with_path = add_image_path(image_path, return_df)
    excel_file = f"{save_dir}/products_with_images.xlsx"
    df_with_path.to_excel(excel_file, index=False, engine='openpyxl')
    wb = load_workbook(excel_file)
    ws = wb.active

    img_width = 5.0
    img_height = 5.0

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        if not df_with_path.loc[idx - 1, "image_path"]:
            continue
        img_path = f"{save_dir}/" + df_with_path.loc[idx - 1, "image_path"]
        
        if os.path.exists(img_path):
            img = Image(img_path)
            img.width = img_width * 18
            img.height = img_height * 18
            cell = f'B{idx+1}'
            ws.add_image(img, cell)

            ws.row_dimensions[idx + 1].height = img_height * 14
            ws.column_dimensions['B'].width = img_width * 2.5
        else:
            print(f"Error image do not exist - {img_path}")

    wb.save(excel_file)
    print(f"finish：{excel_file}")